import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from datetime import datetime
import os
import threading
import schedule
import time
import glob
import requests
from bs4 import BeautifulSoup
import urllib3
import re
import urllib.parse
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
import winreg as reg

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="pandas")

# ── 📦 PORTABLE APPDATA RESOLVER ──
import sys
if getattr(sys, 'frozen', False):
    # Running as PyInstaller .exe
    _BASE_DIR = os.path.dirname(sys.executable)
else:
    # Running as standard .py
    _BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Load secure credentials explicitly from the app folder
load_dotenv(os.path.join(_BASE_DIR, ".env"))

# ==========================================
# ⚙️ MASTER CONFIGURATION
# ==========================================
PORTAL_URL = "http://122.186.209.30:8068/NCC/Sitapur/Sign-In-Users.php"
MY_USER = os.getenv("PORTAL_USER")
MY_PASS = os.getenv("PORTAL_PASS")
MY_DISTRICT = os.getenv("PORTAL_DISTRICT", "Sitapur")

CHROME_DATA_DIR = os.path.join(_BASE_DIR, "Nexus_Chrome_Profile")
CONTACT_FILE = os.path.join(_BASE_DIR, "whatsapp_contacts.txt")

# Bright "Arctic Ice" Theme
CLR_BG = "#f8f9fa"         # Light crisp gray
CLR_SIDEBAR = "#ffffff"    # Pure white
CLR_CARD = "#ffffff"       # Pure white cards
CLR_BORDER = "#d1d5db"     # Soft gray borders
CLR_CYAN = "#0ea5e9"       # Sky Blue
CLR_GREEN = "#10b981"      # Emerald Green
CLR_GOLD = "#f59e0b"       # Bright Amber
CLR_TEXT = "#111827"       # Very dark gray (almost black)
CLR_DIM = "#6b7280"        # Slate gray for secondary text

class NexusSyncPro(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("NEXUS SYNC | Enterprise Suite v9.0 (Prod)")
        self.geometry("1600x950")
        self.configure(fg_color=CLR_BG)
        
        self.service_active = tk.BooleanVar(value=True)
        self.contacts = self.load_contacts()
        self.browser_lock = threading.Lock()
        self._jjm_cache = {"count": "0", "timestamp": 0.0}
        
        self.setup_ui()
        
        os.makedirs(CHROME_DATA_DIR, exist_ok=True)
        self.today_str = datetime.today().strftime("%d-%m-%Y")

        # ── INITIALIZE WORKSPACE (Auto-selects last location) ──
        self.watch_folder = self._select_workspace_folder()

        self.safe_log_update("[SYS] System Architecture v9.0 (Production Release) Initialized.")
        self.safe_log_update(f"[SYS] Daily data directory mapped: {self.watch_folder}")
        if os.listdir(self.watch_folder):
            self.safe_log_update(f"[SYS] Existing files detected in today's folder — reusing workspace.")
        
        self._register_startup()
        
        threading.Thread(target=self.run_scheduler, daemon=True).start()
        threading.Thread(target=self.startup_check, daemon=True).start()

    def _select_workspace_folder(self, force_prompt=False):
        """Pick a base folder. Auto-selects if config exists unless force_prompt is True."""
        config_path = os.path.join(os.getcwd(), ".nexus_workspace_path")
        last_dir = None
        if os.path.exists(config_path):
            with open(config_path, "r") as fp:
                last_dir = fp.read().strip() or None

        if not force_prompt and last_dir and os.path.exists(last_dir):
            chosen = last_dir
        else:
            self.safe_log_update("[SYS] Requesting workspace folder selection...")
            self.update()
            chosen = filedialog.askdirectory(
                title="Select Base Workspace Folder for NEXUS SYNC Data",
                initialdir=last_dir or os.path.expanduser("~"),
                mustexist=True,
            )

        if not chosen:
            if last_dir: 
                chosen = last_dir 
            else:
                self.safe_log_update("[SYS] No folder chosen — using default workspace.")
                chosen = os.path.join(os.getcwd(), "JJM_Daily_Data")

        # Persist the base folder choice
        with open(config_path, "w") as fp:
            fp.write(chosen)

        dated_folder = os.path.join(chosen, self.today_str)
        os.makedirs(dated_folder, exist_ok=True)
        return dated_folder

    def manual_change_workspace(self):
        new_folder = self._select_workspace_folder(force_prompt=True)
        if new_folder:
            self.watch_folder = new_folder
            self.safe_log_update(f"[SYS] Workspace switched: {self.watch_folder}")
            self.rebuild_history_log()
            # Run a fresh analysis on the new folder context
            files = glob.glob(os.path.join(self.watch_folder, '*.xlsx'))
            raw_files = [f for f in files if not os.path.basename(f).startswith("Final_Daily_Report")]
            if raw_files:
                self.analyze_data(max(raw_files, key=os.path.getctime))

    def _register_startup(self):
        """Register the app in the Windows Registry (Boot) AND Task Scheduler (8 AM Daily)."""
        app_path = sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__)
        if not app_path.endswith(".exe"):
            return # Skip if not compiled

        # 1. Windows startup registry (for boot)
        try:
            key = reg.HKEY_CURRENT_USER
            sub_key = r"Software\Microsoft\Windows\CurrentVersion\Run"
            with reg.OpenKey(key, sub_key, 0, reg.KEY_ALL_ACCESS) as registry_key:
                reg.SetValueEx(registry_key, "NexusSyncPro", 0, reg.REG_SZ, f'"{app_path}"')
            self.safe_log_update("[SYS] Configured automatic startup on Windows boot.")
        except Exception as e:
            self.safe_log_update(f"[SYS] ⚠️ Registry startup fail: {str(e)}")

        # 2. Windows Task Scheduler (for 8 AM Daily Re-open)
        try:
            cmd = f'schtasks /create /tn "NexusSyncPro_DailyOpen" /tr "\\"{app_path}\\"" /sc daily /st 08:00 /f'
            os.system(cmd)
            self.safe_log_update("[SYS] Configured Daily Re-open Task for 08:00 AM.")
        except Exception as e:
            self.safe_log_update(f"[SYS] ⚠️ Task Scheduler fail: {str(e)}")

    def auto_close_app(self):
        """End of day clean shutdown."""
        self.safe_log_update("\n[SYS] 7:00 PM REACHED. System performing scheduled shutdown...")
        self.update()
        time.sleep(3)
        os._exit(0)

    def setup_ui(self):
        # --- SIDEBAR ---
        self.sidebar = ctk.CTkFrame(self, fg_color=CLR_SIDEBAR, width=300, corner_radius=0)
        self.sidebar.pack(side="left", fill="y")
        
        ctk.CTkLabel(self.sidebar, text="NEXUS SYNC", font=("Segoe UI", 28, "bold"), text_color=CLR_CYAN).pack(pady=(40, 10))
        ctk.CTkLabel(self.sidebar, text="PRODUCTION BUILD", font=("Segoe UI", 10, "bold"), text_color=CLR_GOLD).pack(pady=(0, 30))

        ctrl_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        ctrl_frame.pack(fill="x", padx=20, pady=10)
        
        self.service_toggle = ctk.CTkSwitch(ctrl_frame, text="AUTO-PILOT MODE", variable=self.service_active, 
                                           progress_color=CLR_GREEN, font=("Segoe UI", 12, "bold"))
        self.service_toggle.pack(pady=10)
        
        self.pull_btn = ctk.CTkButton(ctrl_frame, text="📥 PULL NEW DATA", fg_color="#f1f5f9", hover_color="#e2e8f0", 
                                     border_width=1, border_color=CLR_CYAN, text_color=CLR_CYAN, font=("Segoe UI", 13, "bold"), height=40, command=self.trigger_manual_pull)
        self.pull_btn.pack(fill="x", pady=5)
        
        self.send_btn = ctk.CTkButton(ctrl_frame, text="📤 BROADCAST REPORT", fg_color=CLR_GREEN, hover_color="#059669", 
                                     text_color="#ffffff", font=("Segoe UI", 13, "bold"), height=40, command=self.trigger_manual_send)
        self.send_btn.pack(fill="x", pady=5)

        self.report_btn = ctk.CTkButton(ctrl_frame, text="📊 GENERATE DAILY REPORT", fg_color=CLR_GOLD, hover_color="#d97706", 
                                     text_color="#ffffff", font=("Segoe UI", 13, "bold"), height=40, command=self.generate_final_report)
        self.report_btn.pack(fill="x", pady=5)

        self.ws_btn = ctk.CTkButton(ctrl_frame, text="📁 CHANGE WORKSPACE", fg_color="transparent", border_width=1, border_color=CLR_BORDER,
                                     text_color=CLR_TEXT, font=("Segoe UI", 13, "bold"), height=40, command=self.manual_change_workspace)
        self.ws_btn.pack(fill="x", pady=5)

        ctk.CTkFrame(self.sidebar, fg_color=CLR_BORDER, height=2).pack(fill="x", padx=30, pady=20)

        ctk.CTkLabel(self.sidebar, text="📒 CONTACT BOOK", font=("Segoe UI", 10, "bold"), text_color=CLR_CYAN).pack(anchor="w", padx=30, pady=(0, 6))

        cb_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        cb_frame.pack(fill="x", padx=20)
        ctk.CTkLabel(cb_frame, text="Name", font=("Segoe UI", 10), text_color="#9599a1").pack(anchor="w")
        self.contact_name_entry = ctk.CTkEntry(cb_frame, placeholder_text="e.g. Gopi Sir", height=34, fg_color=CLR_BG, border_color=CLR_BORDER)
        self.contact_name_entry.pack(fill="x", pady=(2, 6))
        ctk.CTkLabel(cb_frame, text="Phone (with country code)", font=("Segoe UI", 10), text_color="#9599a1").pack(anchor="w")
        self.contact_phone_entry = ctk.CTkEntry(cb_frame, placeholder_text="e.g. 919951092727", height=34, fg_color=CLR_BG, border_color=CLR_BORDER)
        self.contact_phone_entry.pack(fill="x", pady=(2, 8))
        ctk.CTkButton(cb_frame, text="+ Add Contact", command=self.add_contact,
                      fg_color="transparent", border_width=1, border_color=CLR_BORDER,
                      font=("Segoe UI", 12, "bold"), height=34).pack(fill="x")

        self.contact_listbox = tk.Listbox(self.sidebar, bg=CLR_BG, fg=CLR_TEXT, borderwidth=0,
                                          highlightthickness=0, font=("Segoe UI", 11),
                                          selectbackground=CLR_CYAN, selectforeground=CLR_BG)
        self.contact_listbox.pack(fill="both", expand=True, padx=25, pady=10)
        self.refresh_contact_ui()
        ctk.CTkButton(self.sidebar, text="🗑 Remove Selected", text_color="#ff4d4d",
                      fg_color="transparent", command=self.remove_contact).pack(pady=(0, 10))

        # ── DEVELOPER CREDIT ──
        ctk.CTkLabel(self.sidebar, text="DEVELOPED BY: ASHISH KUMAR", font=("Segoe UI", 9, "italic"), text_color=CLR_DIM).pack(side="bottom", pady=20)

        # --- MAIN DISPLAY ---
        self.display = ctk.CTkFrame(self, fg_color="transparent")
        self.display.pack(side="right", fill="both", expand=True, padx=20, pady=20)

        # Top Split Frame (Logs & History)
        self.top_split = ctk.CTkFrame(self.display, fg_color="transparent")
        self.top_split.pack(fill="x", pady=(0, 15))

        # 1. System Log Terminal
        self.log_container = ctk.CTkFrame(self.top_split, fg_color=CLR_CARD, border_width=1, border_color=CLR_BORDER, height=260)
        self.log_container.pack(side="left", fill="both", expand=True, padx=(0, 10))
        self.log_container.pack_propagate(False) 
        ctk.CTkLabel(self.log_container, text="📡 SYSTEM LOG ENGINE", font=("Segoe UI", 12, "bold"), text_color=CLR_CYAN).pack(anchor="w", padx=15, pady=5)
        self.log_terminal = ctk.CTkTextbox(self.log_container, fg_color="#f3f4f6", text_color=CLR_TEXT, font=("Consolas", 12))
        self.log_terminal.pack(fill="both", expand=True, padx=10, pady=10)

        # 2. Daily Mapping History Terminal
        self.history_container = ctk.CTkFrame(self.top_split, fg_color=CLR_CARD, border_width=1, border_color=CLR_BORDER, height=260)
        self.history_container.pack(side="right", fill="both", expand=True, padx=(10, 0))
        self.history_container.pack_propagate(False)
        ctk.CTkLabel(self.history_container, text="⏱️ DAILY MAPPING HISTORY", font=("Segoe UI", 12, "bold"), text_color=CLR_CYAN).pack(anchor="w", padx=15, pady=5)
        self.history_terminal = ctk.CTkTextbox(self.history_container, fg_color="#f3f4f6", text_color=CLR_CYAN, font=("Consolas", 13))
        self.history_terminal.pack(fill="both", expand=True, padx=10, pady=10)

        # Bottom Split Frame (Reports & WhatsApp)
        self.bottom_split = ctk.CTkFrame(self.display, fg_color="transparent")
        self.bottom_split.pack(fill="both", expand=True)

        # 3. Data Analysis Terminal
        self.report_container = ctk.CTkFrame(self.bottom_split, fg_color=CLR_CARD, border_width=1, border_color=CLR_BORDER)
        self.report_container.pack(side="left", fill="both", expand=True, padx=(0, 10))
        ctk.CTkLabel(self.report_container, text="📊 DATA ANALYSIS REPORT", font=("Segoe UI", 12, "bold"), text_color=CLR_GOLD).pack(anchor="w", padx=15, pady=5)
        self.report_terminal = ctk.CTkTextbox(self.report_container, fg_color="#f3f4f6", text_color=CLR_TEXT, font=("Consolas", 13))
        self.report_terminal.pack(fill="both", expand=True, padx=10, pady=10)

        # 4. WhatsApp Preview Terminal
        self.preview_container = ctk.CTkFrame(self.bottom_split, fg_color=CLR_CARD, border_width=1, border_color=CLR_BORDER)
        self.preview_container.pack(side="right", fill="both", expand=True, padx=(10, 0))
        ctk.CTkLabel(self.preview_container, text="📱 WHATSAPP PAYLOAD PREVIEW", font=("Segoe UI", 12, "bold"), text_color=CLR_GREEN).pack(anchor="w", padx=15, pady=5)
        self.preview_terminal = ctk.CTkTextbox(self.preview_container, fg_color="#f3f4f6", text_color=CLR_TEXT, font=("Segoe UI", 13))
        self.preview_terminal.pack(fill="both", expand=True, padx=10, pady=10)

    # --- UI Thread Safety Wrappers ---
    def safe_log_update(self, text):
        self.after(0, lambda: self._log_update(text))

    def _log_update(self, text):
        self.log_terminal.insert("end", f"{text}\n")
        self.log_terminal.see("end")

    def safe_history_update(self, text):
        self.after(0, lambda: self._history_update(text))

    def _history_update(self, text):
        self.history_terminal.insert("end", f"{text}\n")
        self.history_terminal.see("end")

    def safe_report_update(self, report_text, preview_text):
        self.after(0, lambda: self._report_update(report_text, preview_text))

    def _report_update(self, report_text, preview_text):
        self.report_terminal.delete("1.0", "end")
        self.report_terminal.insert("end", report_text)
        self.preview_terminal.delete("1.0", "end")
        self.preview_terminal.insert("end", preview_text)

    # ==================================================
    # 🤖 CORE LOGIC
    # ==================================================
    
    def rebuild_history_log(self):
        self.after(0, lambda: self.history_terminal.delete("1.0", "end"))
        files = glob.glob(os.path.join(self.watch_folder, '*.xlsx'))
        raw_files = [f for f in files if not os.path.basename(f).startswith("Final_Daily_Report")]
        raw_files.sort(key=os.path.getctime)
        
        if raw_files:
            self.safe_history_update(f"[SYS] Fetched {len(raw_files)} data modules from workspace.")
            self.safe_history_update("[SYS] Initializing core engine...\n")
            for f in raw_files:
                try:
                    df = pd.read_excel(f, header=None, nrows=15)
                    header_row = df.notna().sum(axis=1).idxmax()
                    df = pd.read_excel(f, header=header_row)
                    df.columns = [str(c).strip().lower() for c in df.columns]
                    dt_col = next((c for c in df.columns if "date" in c or "time" in c), df.columns[-1])
                    df[dt_col] = pd.to_datetime(df[dt_col], errors="coerce", format="mixed")
                    today = datetime.today().date()
                    
                    synced = len(df[df[dt_col].dt.date == today])
                    not_synced = len(df[df[dt_col].dt.date != today])
                    
                    c_time = datetime.fromtimestamp(os.path.getctime(f))
                    timestamp = c_time.strftime("%b %d - %I:%M %p")
                    self.safe_history_update(f"[+] Mapped: {timestamp} | Sync: {synced:03d} | Unsync: {not_synced:03d}")
                except Exception:
                    pass
        else:
             self.safe_history_update("[SYS] Workspace empty. Awaiting first download...")

    def auto_fetch_jjm_count(self):
        """Lightweight HTTP fetch — cached for 2 minutes."""
        if (time.time() - self._jjm_cache["timestamp"]) < 120:
            return self._jjm_cache["count"]

        self.safe_log_update("> [WEB] Connecting to JJM UP Portal...")
        url = "https://jjm.up.gov.in/site/OS_AutomationIntegration_DistrictWise"
        try:
            hdrs = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'}
            response = requests.get(url, headers=hdrs, timeout=15, verify=False)

            if response.status_code != 200:
                self.safe_log_update(f"❌ [WEB] Website rejected connection (HTTP {response.status_code})")
                return "0"

            self.safe_log_update("> [WEB] Connection successful! Reading data table...")
            soup = BeautifulSoup(response.text, 'html.parser')
            tables = soup.find_all('table')

            if not tables:
                self.safe_log_update("❌ [WEB] Error: Table not found on the page.")
                return "0"

            current_district = ""
            for row in tables[0].find_all('tr'):
                cells = row.find_all(['td', 'th'])
                texts = [c.text.strip() for c in cells]

                # Track district across rows (Sitapur & NCC may be on different rows)
                for text in texts:
                    if "Sitapur" in text:
                        current_district = "Sitapur"
                        break

                if current_district == "Sitapur":
                    agency_idx = -1
                    for i, t in enumerate(texts):
                        if "NCC" in t.upper():
                            agency_idx = i
                            break

                    if agency_idx != -1:
                        self.safe_log_update("🎯 [WEB] Found Sitapur → NCC Ltd!")
                        try:
                            target_val = texts[agency_idx + 10]
                            clean_number = re.sub(r'\D', '', target_val)
                            self.safe_log_update(f"⮑ [WEB] Live Connected Scheme Count Fetched: {clean_number}")
                            self._jjm_cache = {"count": clean_number, "timestamp": time.time()}
                            return clean_number
                        except IndexError:
                            self.safe_log_update("❌ [WEB] Table structure mismatch — column offset wrong.")
                            return "0"

            self.safe_log_update("❌ [WEB] Sitapur NCC Ltd data not found in table.")
            return "0"
        except Exception as e:
            self.safe_log_update(f"❌ [WEB] Connection failed: {str(e)}")
            return "0"

    def robot_portal_download(self, force=False):
        with self.browser_lock:
            self.safe_log_update("\n[SYS] Initiating portal data pull...")

            # ── ANTI-DUPLICATE DEBOUNCE: Skip if we just downloaded data ──
            existing_files = glob.glob(os.path.join(self.watch_folder, '*.xlsx'))
            if not force and existing_files:
                latest_file = max(existing_files, key=os.path.getctime)
                if (time.time() - os.path.getctime(latest_file)) < 60: # 1 minute
                    self.safe_log_update("⚠️ System pulled data less than a minute ago. Using recent data...")
                    self.analyze_data(latest_file)
                    return

            driver = None
            downloaded_file = None
            try:
                if not MY_USER or not MY_PASS:
                    self.safe_log_update("❌ Error: Credentials missing in .env file.")
                    return

                # Snapshot files before download so we can detect the NEW one
                before_files = set(glob.glob(os.path.join(self.watch_folder, '*.xlsx')))

                self.safe_log_update("> [PORTAL] Launching browser & navigating to login page...")
                driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=self.get_armored_options())
                driver.get(PORTAL_URL)
                wait = WebDriverWait(driver, 25)

                self.safe_log_update("> [PORTAL] Waiting for login form...")
                wait.until(EC.presence_of_element_located((By.ID, "username"))).send_keys(MY_USER)
                driver.find_element(By.ID, "password").send_keys(MY_PASS)

                self.safe_log_update(f"> [PORTAL] Selecting district: {MY_DISTRICT}...")
                dropdown = Select(driver.find_element(By.ID, "selectDistt"))
                for opt in dropdown.options:
                    if MY_DISTRICT.lower() in opt.text.lower():
                        dropdown.select_by_visible_text(opt.text)
                        break

                driver.find_element(By.NAME, "login").click()
                self.safe_log_update("> [PORTAL] Login submitted. Waiting for dashboard to load...")
                time.sleep(6)

                self.safe_log_update("> [PORTAL] Dashboard loaded. Triggering Excel export...")
                btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'buttons-excel')]")))
                driver.execute_script("arguments[0].click();", btn)
                self.safe_log_update("> [PORTAL] Export clicked. Polling for downloaded file (max 60s)...")

                # Poll until a NEW .xlsx appears in watch_folder (up to 60 seconds)
                timeout, elapsed = 60, 0
                while elapsed < timeout:
                    time.sleep(2); elapsed += 2
                    current_files = set(glob.glob(os.path.join(self.watch_folder, '*.xlsx')))
                    new_files = [f for f in current_files - before_files
                                 if not os.path.basename(f).startswith("Final_Daily_Report")]
                    if new_files:
                        df_tmp = max(new_files, key=os.path.getctime)
                        try:
                            # Verify no temporary chrome downloads are active & file has data
                            if os.path.getsize(df_tmp) > 0 and not glob.glob(os.path.join(self.watch_folder, '*.crdownload')):
                                time.sleep(1) # Extra buffer for I/O flush
                                downloaded_file = df_tmp
                                self.safe_log_update(f"✅ [PORTAL] File captured: {os.path.basename(downloaded_file)}")
                                break
                        except Exception:
                            pass

                if not downloaded_file:
                    self.safe_log_update("❌ [PORTAL] Download timed out — no new file detected.")

            except Exception as e:
                self.safe_log_update(f"❌ Portal Error: {str(e)}")
            finally:
                # ⚠️ CRITICAL: Close portal browser BEFORE opening JJM browser
                # Both use the same Chrome profile — only one can be open at a time
                if driver:
                    driver.quit()
                    self.safe_log_update("> [PORTAL] Browser closed.")

            # Analyze AFTER browser is fully closed (avoids Chrome profile conflict)
            if downloaded_file:
                self.rebuild_history_log()
                self.analyze_data(downloaded_file)

    def analyze_data(self, latest_file_path):
        try:
            jjm_count = self.auto_fetch_jjm_count()
            if not jjm_count: jjm_count = "0"

            df = pd.read_excel(latest_file_path, header=None, nrows=15)
            header_row = df.notna().sum(axis=1).idxmax()
            df = pd.read_excel(latest_file_path, header=header_row)
            df.columns = [str(c).strip().lower() for c in df.columns]
            
            gp_col = next((c for c in df.columns if "gp" in c or "panchayat" in c or "name" in c), df.columns[1])
            dt_col = next((c for c in df.columns if "date" in c or "time" in c), df.columns[-1])
            
            df[dt_col] = pd.to_datetime(df[dt_col], errors="coerce", format="mixed")
            today = datetime.today().date()
            
            synced = df[df[dt_col].dt.date == today]
            not_synced = df[df[dt_col].dt.date != today]
            
            # Identify Newly Added GPs by comparing against the FIRST file of today
            files = glob.glob(os.path.join(self.watch_folder, '*.xlsx'))
            raw_files = [f for f in files if not os.path.basename(f).startswith("Final_Daily_Report")]
            raw_files.sort(key=os.path.getctime)
            
            daily_new_gps = []
            if len(raw_files) > 1:
                first_file = raw_files[0]
                if first_file != latest_file_path:
                    try:
                        df_old = pd.read_excel(first_file, header=None, nrows=15)
                        h_row = df_old.notna().sum(axis=1).idxmax()
                        df_old = pd.read_excel(first_file, header=h_row)
                        df_old.columns = [str(c).strip().lower() for c in df_old.columns]
                        old_gp_col = next((c for c in df_old.columns if "gp" in c or "panchayat" in c or "name" in c), df_old.columns[1])
                        old_gps_baseline = set(df_old[old_gp_col].dropna().astype(str).unique())
                        
                        current_gps = set(df[gp_col].dropna().astype(str).unique())
                        daily_new_gps = list(current_gps - old_gps_baseline)
                    except: pass
            
            new_gp_count = len(daily_new_gps)
            
            new_gp_text = ""
            if new_gp_count > 0:
                joined_names = ", ".join(sorted(daily_new_gps))
                new_gp_text = f"\nNewly Added Schemes- {new_gp_count} ({joined_names})"
            
            report = f"JJM PORTAL: {jjm_count}\nSCADA TOTAL: {len(df)}\nSYNCED: {len(synced)}\nUNSYNCED: {len(not_synced)}\nNEW ADDED: {new_gp_count}\n\n"
            report += "✅ SYNCED LIST:\n" + ", ".join(synced[gp_col].astype(str).tolist()) + "\n\n"
            report += "❌ NOT SYNCED LIST:\n" + ", ".join(not_synced[gp_col].astype(str).tolist()) + "\n\n"
            if new_gp_count > 0:
                report += "⭐ NEWLY ADDED LIST:\n" + joined_names
            
            greet = "Good Evening Sir," if datetime.now().hour >= 16 else ("Good Afternoon Sir," if datetime.now().hour >= 12 else "Good Morning Sir,")
            preview = f"Date: {datetime.today().strftime('%d.%m.%Y')}\n\n{greet}\nNo of schemes connected with SCADA- {len(synced)}\nNo of schemes Lives in JJM Portal- {jjm_count}{new_gp_text}\n\nToday Schemes listed in SCADA- {len(df)}."
            
            self.last_analysis_msg = preview
            self.safe_report_update(report, preview)
            
            timestamp = datetime.now().strftime("%b %d - %I:%M %p")
            history_line = f"[+] Mapped: {timestamp} | Sync: {len(synced):03d} | Unsync: {len(not_synced):03d}"
            self.safe_history_update(history_line)
            self.safe_log_update("[SYS] Data structured and ready for export.")
            
        except Exception as e: self.safe_log_update(f"❌ Analysis Fail: {str(e)}")

    def generate_final_report(self):
        self.safe_log_update("\n[SYS] Generating End of Day Final Report...")
        files = glob.glob(os.path.join(self.watch_folder, '*.xlsx'))
        raw_files = [f for f in files if not os.path.basename(f).startswith("Final_Daily_Report")]
        
        if not raw_files:
            self.safe_log_update("❌ No raw data files found for today.")
            return

        def detect_hour(filepath):
            filename = os.path.basename(filepath)
            date_match = re.search(r'(\d{8})_(\d{4})', filename)
            if date_match:
                try:
                    dt_obj = datetime.strptime(f"{date_match.group(1)}_{date_match.group(2)}", "%Y%m%d_%H%M")
                    return dt_obj, dt_obj.strftime("%b %d - %I:%M %p")
                except ValueError: pass

            time_match = re.search(r'_(\d{4})|(\d{4})', filename)
            if time_match:
                t = time_match.group(1) or time_match.group(2)
                try:
                    time_obj = datetime.strptime(t, "%H%M").time()
                    file_date = datetime.fromtimestamp(os.path.getmtime(filepath)).date()
                    dt_obj = datetime.combine(file_date, time_obj)
                    return dt_obj, dt_obj.strftime("%b %d - %I:%M %p")
                except ValueError: pass
                    
            try:
                timestamp = os.path.getmtime(filepath)
                dt_obj = datetime.fromtimestamp(timestamp)
                return dt_obj, dt_obj.strftime("%b %d - %I:%M %p")
            except Exception: pass
            return None, None
            
        def sort_by_time(filepath):
            dt_obj, _ = detect_hour(filepath)
            return dt_obj if dt_obj else datetime.max 

        raw_files.sort(key=sort_by_time)

        try:
            # 1. READ ALL FILES AND BUILD HOURLY DATA
            hourly_data = []
            today = datetime.today().date()
            previous_hour_gps = set()

            # Find actual columns from the first file
            df_first = pd.read_excel(raw_files[0], header=None, nrows=15)
            header_row_first = df_first.notna().sum(axis=1).idxmax()
            if isinstance(header_row_first, str): header_row_first = 0
            df_first = pd.read_excel(raw_files[0], header=header_row_first)
            df_first.columns = [str(c).strip() for c in df_first.columns]
            headers = list(df_first.columns)
            
            actual_sno = next((c for c in headers if "s no" in str(c).lower() or "s.no" in str(c).lower() or "sl" in str(c).lower() or "serial" in str(c).lower()), headers[0])
            rem_h = [c for c in headers if c != actual_sno]
            actual_primary = next((c for c in rem_h if "name" in str(c).lower() or "gp" in str(c).lower() or "panchayat" in str(c).lower()), rem_h[0] if rem_h else headers[0])
            rem_h2 = [c for c in rem_h if c != actual_primary]
            actual_dt = next((c for c in rem_h2 if "date" in str(c).lower() or "time" in str(c).lower()), rem_h2[0] if rem_h2 else headers[-1])

            for idx, f in enumerate(raw_files):
                hour, display = detect_hour(f)
                if hour is None: continue
                
                # Smart read
                df_raw = pd.read_excel(f, header=None, nrows=15)
                header_row = 0
                for i, row in df_raw.iterrows():
                    if row.notna().sum() >= 3:
                        header_row = i
                        break
                df = pd.read_excel(f, header=header_row)
                df.columns = [str(c).strip() for c in df.columns]
                df = df.loc[:, ~df.columns.str.contains('^Unnamed')]

                primary_col = actual_primary if actual_primary in df.columns else df.columns[0]
                dt_col = actual_dt if actual_dt in df.columns else df.columns[1]

                current_hour_gps = set(df[primary_col].dropna().astype(str).unique())
                previous_hour_gps = current_hour_gps

                df[dt_col] = pd.to_datetime(df[dt_col], errors="coerce", format="mixed")
                syncing = df[df[dt_col].dt.date == today]
                unsync = df[df[dt_col].dt.date != today]

                hourly_data.append({"hour": hour, "display": display, "sync": syncing, "unsync": unsync, "raw": df})

            if not hourly_data:
                self.safe_log_update("❌ No readable data in today's files.")
                return

            # 2. EXPORT MATRIX TO EXCEL
            final_filename = f"Final_Daily_Report_{self.today_str}.xlsx"
            final_path = os.path.join(self.watch_folder, final_filename)

            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

            wb = Workbook()
            ws = wb.active
            ws.freeze_panes = "A2"
            
            sync_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            unsync_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            new_gp_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid") 
            header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") 
            total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            border = Border(left=Side(style='thin', color="CBD5E1"), right=Side(style='thin', color="CBD5E1"), top=Side(style='thin', color="CBD5E1"), bottom=Side(style='thin', color="CBD5E1"))

            extra_cols = []
            master_df = pd.concat([h["raw"] for h in hourly_data]).drop_duplicates(subset=[actual_primary])
            def get_priority(pk):
                for h in hourly_data:
                    if pk in h["sync"][actual_primary].values: return 0
                return 1
            master_df['_priority'] = master_df[actual_primary].apply(get_priority)
            master_df = master_df.sort_values(by=['_priority', actual_primary])

            all_headers = [actual_sno, actual_primary] + extra_cols + [h["display"] for h in hourly_data]
            for c, text in enumerate(all_headers, 1):
                cell = ws.cell(row=1, column=c, value=text)
                cell.font, cell.fill, cell.border = Font(bold=True, color="FFFFFF", size=11), header_fill, border
                cell.alignment = Alignment(horizontal="center")

            new_gp_counts = {i: 0 for i in range(len(hourly_data))}
            r_idx = 2
            
            for _, row_data in master_df.iterrows():
                ws.cell(row=r_idx, column=1, value=row_data.get(actual_sno, "-")).border = border
                ws.cell(row=r_idx, column=2, value=row_data.get(actual_primary, "-")).border = border
                
                for i, col in enumerate(extra_cols, 3):
                    ws.cell(row=r_idx, column=i, value=row_data.get(col, "-")).border = border

                for i, h_data in enumerate(hourly_data, len(extra_cols) + 3):
                    cell = ws.cell(row=r_idx, column=i)
                    cell.border = border
                    list_idx = i - (len(extra_cols) + 3) 

                    pk = row_data[actual_primary]
                    s_match = h_data["sync"][h_data["sync"][actual_primary] == pk]
                    u_match = h_data["unsync"][h_data["unsync"][actual_primary] == pk]

                    is_newly_added = False
                    if list_idx > 0: 
                        prev_h_data = hourly_data[list_idx - 1]
                        in_current = (not s_match.empty) or (not u_match.empty)
                        in_prev = pk in prev_h_data["raw"][actual_primary].values
                        if in_current and not in_prev:
                            is_newly_added = True
                            new_gp_counts[list_idx] += 1 

                    if not s_match.empty:
                        val = s_match.iloc[0][actual_dt]
                        cell.value = val.strftime("%Y-%m-%d %H:%M") if pd.notnull(val) else "N/A"
                        cell.fill = new_gp_fill if is_newly_added else sync_fill
                    elif not u_match.empty:
                        val = u_match.iloc[0][actual_dt]
                        cell.value = val.strftime("%Y-%m-%d %H:%M") if pd.notnull(val) else "N/A"
                        cell.fill = new_gp_fill if is_newly_added else unsync_fill
                    else:
                        cell.value = "-"
                r_idx += 1

            r_idx += 1
            ws.cell(row=r_idx, column=2, value="SUCCESS COUNT (TODAY)").font = Font(bold=True)
            ws.cell(row=r_idx+1, column=2, value="STALE COUNT (OLD)").font = Font(bold=True)
            ws.cell(row=r_idx+2, column=2, value="NEW GP COUNT (ADDED)").font = Font(bold=True) 

            for i, h_data in enumerate(hourly_data, len(extra_cols) + 3):
                list_idx = i - (len(extra_cols) + 3)
                s_c = ws.cell(row=r_idx, column=i, value=len(h_data["sync"]))
                u_c = ws.cell(row=r_idx+1, column=i, value=len(h_data["unsync"]))
                n_c = ws.cell(row=r_idx+2, column=i, value=new_gp_counts[list_idx]) 
                for c in [s_c, u_c, n_c]:
                    c.fill, c.font, c.border, c.alignment = total_fill, Font(bold=True), border, Alignment(horizontal="center")

            wb.save(final_path)
            self.safe_log_update(f"✅ Final Matrix Report Saved: {final_path}")

        except Exception as e:
            self.safe_log_update(f"❌ Failed to generate matrix report: {str(e)}")

    # ==================================================
    # 📱 WHATSAPP ENGINE (RESTORED MANUAL METHOD)
    # ==================================================
    
    def robot_whatsapp_blast(self):
        with self.browser_lock:
            self.safe_log_update("\n--- 📱 WHATSAPP BROADCAST ENGINE ---")
            if not self.contacts or not self.last_analysis_msg:
                self.safe_log_update("⚠️ Missing contacts or report. Pull data first.")
                return

            driver = None
            try:
                driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=self.get_armored_options(False))
                driver.get("https://web.whatsapp.com")
                self.safe_log_update("[WA] Browser opened. Scan QR if needed. Waiting 20s for interface...")
                time.sleep(20)

                for contact in self.contacts:
                    name  = contact.get('name', '')
                    phone = contact.get('phone', '')
                    self.safe_log_update(f"\n-> Sending to: {name} ({phone})")
                    try:
                        # Always use direct URL — 100% reliable since all contacts have a phone
                        driver.get(f"https://web.whatsapp.com/send?phone={phone}")
                        time.sleep(7)

                        # ── FIND MESSAGE BOX ──
                        msg_box = None
                        for xpath in [
                            '//div[@contenteditable="true"][@data-tab="10"]',
                            '//div[@title="Type a message"]',
                            '//div[contains(@aria-label,"Type a message")][@contenteditable="true"]',
                            '//div[contains(@aria-label,"message")][@contenteditable="true"]',
                        ]:
                            try:
                                msg_box = WebDriverWait(driver, 12).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                                break
                            except: continue

                        if not msg_box:
                            self.safe_log_update(f"   ❌ Message box not found. Is {phone} on WhatsApp?")
                            continue

                        msg_box.click(); time.sleep(0.5)

                        # ── TYPE MESSAGE line by line (Shift+Enter between lines) ──
                        self.safe_log_update("   [WA] Injecting payload...")
                        for line in self.last_analysis_msg.split('\n'):
                            msg_box.send_keys(line)
                            ActionChains(driver).key_down(Keys.SHIFT).send_keys(Keys.ENTER).key_up(Keys.SHIFT).perform()
                            time.sleep(0.08)

                        time.sleep(0.5)
                        msg_box.send_keys(Keys.ENTER)
                        self.safe_log_update(f"   ✅ Sent to {name}!")
                        time.sleep(3)

                    except Exception as e:
                        self.safe_log_update(f"   ❌ Failed for {name}: {str(e)}")
                        continue

            except Exception as e:
                self.safe_log_update(f"❌ WhatsApp Engine Error: {str(e)}")
            finally:
                if driver:
                    driver.quit()
                    self.safe_log_update("\n[WA] Session closed.")

            # ── AUTO-EXPORT: Generate Final Report after broadcast ──
            self.safe_log_update("\n[SYS] Auto-generating Final Daily Report post-broadcast...")
            self.generate_final_report()

    # ==================================================
    # ⚙️ UTILITIES & BOOT
    # ==================================================
    def get_armored_options(self, is_download=True):
        o = webdriver.ChromeOptions()
        o.add_argument(f"user-data-dir={CHROME_DATA_DIR}")
        o.add_argument("--no-sandbox")
        o.add_argument("--disable-dev-shm-usage")
        o.add_argument("--disable-gpu")
        o.add_argument("--remote-allow-origins=*")
        o.add_argument("--ignore-certificate-errors")
        o.add_argument("--allow-running-insecure-content")
        o.add_argument("--disable-features=InsecureDownloadWarnings")
        o.add_argument("--unsafely-treat-insecure-origin-as-secure=http://122.186.209.30:8068")
        if is_download: 
            prefs = {
                "download.default_directory": os.path.abspath(self.watch_folder), 
                "download.prompt_for_download": False, 
                "download.directory_upgrade": True,
                "safebrowsing.enabled": False,
                "safebrowsing.disable_download_protection": True,
                "safebrowsing_for_trusted_sources_enabled": False,
                "profile.default_content_setting_values.automatic_downloads": 1,
                "profile.default_content_settings.popups": 0
            }
            o.add_experimental_option("prefs", prefs)
        return o

    def load_contacts(self):
        """Load contacts as list of dicts {name, phone}. Supports old plain-number format."""
        contacts = []
        if os.path.exists(CONTACT_FILE):
            with open(CONTACT_FILE, "r") as f:
                for line in f.readlines():
                    line = line.strip()
                    if not line: continue
                    if "|" in line:
                        parts = line.split("|", 1)
                        contacts.append({"name": parts[0].strip(), "phone": parts[1].strip()})
                    else:
                        # Backward compat: plain number entry
                        contacts.append({"name": line, "phone": re.sub(r'\D', '', line)})
        return contacts

    def save_contacts(self):
        with open(CONTACT_FILE, "w") as f:
            for c in self.contacts:
                f.write(f"{c['name']}|{c['phone']}\n")

    def add_contact(self):
        name  = self.contact_name_entry.get().strip()
        phone = re.sub(r'\D', '', self.contact_phone_entry.get().strip())
        if not name or not phone:
            return
        # Prevent duplicate phones
        if any(c['phone'] == phone for c in self.contacts):
            return
        self.contacts.append({"name": name, "phone": phone})
        self.save_contacts()
        self.refresh_contact_ui()
        self.contact_name_entry.delete(0, 'end')
        self.contact_phone_entry.delete(0, 'end')

    def remove_contact(self):
        sel = self.contact_listbox.curselection()
        if sel:
            del self.contacts[sel[0]]
            self.save_contacts()
            self.refresh_contact_ui()

    def show_robot_alert(self, task_name):
        """Centered, self-dismissing warning to notify user of robot takeover."""
        def create_popup():
            popup = ctk.CTkToplevel(self)
            popup.attributes("-topmost", True)
            popup.overrideredirect(True)
            
            # Center on screen
            sw = self.winfo_screenwidth()
            sh = self.winfo_screenheight()
            w, h = 450, 150
            x, y = (sw // 2) - (w // 2), (sh // 2) - (h // 2)
            popup.geometry(f"{w}x{h}+{x}+{y}")
            
            # Minimalist 'Glass' aesthetics
            frame = ctk.CTkFrame(popup, fg_color="#1e293b", border_width=2, border_color=CLR_GOLD)
            frame.pack(fill="both", expand=True)
            
            ctk.CTkLabel(frame, text="⚠️ NEXUS ROBOT ACTIVATING", font=("Segoe UI", 16, "bold"), text_color=CLR_GOLD).pack(pady=(20, 5))
            ctk.CTkLabel(frame, text=f"System will start: {task_name}", font=("Segoe UI", 12), text_color="#cbd5e1").pack()
            ctk.CTkLabel(frame, text="Please STOP using your system now.", font=("Segoe UI", 11, "italic"), text_color="#94a3b8").pack(pady=10)
            
            self.after(3000, popup.destroy)
        
        self.after(0, create_popup)

    def refresh_contact_ui(self):
        self.contact_listbox.delete(0, 'end')
        for c in self.contacts:
            self.contact_listbox.insert('end', f"  {c['name']}  •  {c['phone']}")

    def trigger_manual_pull(self): threading.Thread(target=self.robot_portal_download, kwargs={'force': True}, daemon=True).start()
    def trigger_manual_send(self): threading.Thread(target=self.robot_whatsapp_blast, daemon=True).start()
    
    def startup_check(self):
        self.rebuild_history_log()
        time.sleep(1)
        files = glob.glob(os.path.join(self.watch_folder, '*.xlsx'))
        raw_files = [f for f in files if not os.path.basename(f).startswith("Final_Daily_Report")]
        
        if raw_files:
            latest_file = max(raw_files, key=os.path.getctime)
            latest_time = datetime.fromtimestamp(os.path.getmtime(latest_file))
            
            if latest_time.hour != datetime.now().hour:
                self.safe_log_update(f"[SYS] No data found for current hour ({datetime.now().hour}:00). Auto-pulling...")
                self.show_robot_alert("INITIAL STARTUP SYNC")
                time.sleep(20)
                self.robot_portal_download()
            else:
                self.analyze_data(latest_file)
        else:
            self.safe_log_update("[SYS] No data found for today. Auto-pulling from portal...")
            self.show_robot_alert("FIRST DAILY SYNC")
            time.sleep(20)
            self.robot_portal_download()

    def run_scheduler(self):
        def scheduled_pull():
            self.show_robot_alert("HOURLY DATA REFRESH")
            time.sleep(20)
            self.robot_portal_download()

        def scheduled_blast():
            self.show_robot_alert("EOD WHATSAPP BROADCAST")
            time.sleep(20)
            self.robot_whatsapp_blast()

        for h in range(8, 20): schedule.every().day.at(f"{h:02d}:00").do(self.safe_execute, scheduled_pull)
        schedule.every().day.at("18:05").do(self.safe_execute, scheduled_blast)
        schedule.every().day.at("19:00").do(self.auto_close_app)

        while True:
            schedule.run_pending()
            
            # ── HOURLY DRIFT WATCHDOG ──
            # If it's midway through the hour and we still haven't got data (e.g. net was down)
            if self.service_active.get():
                now = datetime.now()
                if 8 <= now.hour < 19:
                    files = glob.glob(os.path.join(self.watch_folder, '*.xlsx'))
                    raw_files = [f for f in files if not os.path.basename(f).startswith("Final_Daily_Report")]
                    has_current_hour = False
                    for f in raw_files:
                        if datetime.fromtimestamp(os.path.getmtime(f)).hour == now.hour:
                            has_current_hour = True
                            break
                    
                    if not has_current_hour:
                        # Only retry after 10 mins past the hour to allow standard scheduler some buffer
                        if now.minute >= 10: 
                            self.safe_log_update(f"[SYS] Watchdog: Missing data for {now.hour}:00. Attempting recovery pull...")
                            threading.Thread(target=self.robot_portal_download, daemon=True).start()

            time.sleep(60) # Heartbeat: 60 seconds
            
    def safe_execute(self, func):
        """Universal exception wrapper for scheduled tasks."""
        try:
            if not self.service_active.get():
                self.safe_log_update("[SYS] Auto-Pilot is OFF. Skipping scheduled task.")
                return
            func()
        except Exception as e:
            self.safe_log_update(f"[ERR] Scheduler breakdown: {str(e)}")

if __name__ == "__main__":
    app = NexusSyncPro()
    app.mainloop()